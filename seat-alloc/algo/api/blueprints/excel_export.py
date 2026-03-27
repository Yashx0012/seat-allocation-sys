# Excel export endpoint.
# Single xlsx download with:
#   - "Summary" sheet  : plan inputs + room configs + batch breakdown
#   - "Room_<name>" sheet per room : physical seating grid (raw_matrix) + full student list
#
# No metadata section — only inputs-derived data is used.

import logging
from io import BytesIO
from datetime import datetime

from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from algo.services.auth_service import token_required
from algo.core.cache.cache_manager import CacheManager

excel_export_bp = Blueprint('excel_export', __name__, url_prefix='/api')
logger = logging.getLogger(__name__)

_cache_mgr = CacheManager()

# ─────────────────────────────  style helpers  ────────────────────────────────

def _fill(hex_colour: str) -> PatternFill:
    h = hex_colour.lstrip('#')
    return PatternFill(start_color=h, end_color=h, fill_type='solid')


def _font(bold=False, colour='000000', size=10) -> Font:
    return Font(bold=bold, color=colour.lstrip('#'), size=size)


def _align(h='left', v='center', wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


_THIN_SIDE  = Side(style='thin',   color='BBBBBB')
_BORDER     = Border(left=_THIN_SIDE, right=_THIN_SIDE,
                     top=_THIN_SIDE,  bottom=_THIN_SIDE)
_MED_SIDE   = Side(style='medium', color='555555')
_MED_BORDER = Border(left=_MED_SIDE, right=_MED_SIDE,
                     top=_MED_SIDE,  bottom=_MED_SIDE)

# Colour palette (all hex, no leading #)
_C = {
    'sheet_header_bg': '1F4E79',   # Dark navy   - sheet section titles
    'sheet_header_fg': 'FFFFFF',
    'col_header_bg':   '2E75B6',   # Mid blue    - column header rows
    'col_header_fg':   'FFFFFF',
    'unallocated':     'E9ECEF',   # Light grey  - empty seats
    'broken':          'FF8080',   # Red         - broken seats
    'row_even':        'F0F4FA',   # Ice blue    - alternating table rows
    'row_odd':         'FFFFFF',
    'total_bg':        'FFF3CD',   # Amber       - totals row
    'section_bg':      'DEEAF1',   # Pale blue   - config line
}

_ALPHA = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'


# ─────────────────────────  shared cell writers  ──────────────────────────────

def _auto_width(ws, padding=2, cap=55):
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        best       = 0
        col_letter = None
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            try:
                best = max(best, len(str(cell.value)) if cell.value is not None else 0)
            except Exception:
                pass
        if col_letter:
            ws.column_dimensions[col_letter].width = min(best + padding, cap)


def _section_title(ws, row, text, ncols=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font      = _font(bold=True, colour=_C['sheet_header_fg'], size=11)
    cell.fill      = _fill(_C['sheet_header_bg'])
    cell.alignment = _align(h='left')
    cell.border    = _MED_BORDER


def _col_header_row(ws, row, labels):
    for ci, label in enumerate(labels, 1):
        c = ws.cell(row=row, column=ci, value=label)
        c.font      = _font(bold=True, colour=_C['col_header_fg'])
        c.fill      = _fill(_C['col_header_bg'])
        c.alignment = _align(h='center')
        c.border    = _BORDER


def _data_cell(ws, row, col, value, bg=None, bold=False, h='left', wrap=False):
    bg = bg or (_C['row_even'] if row % 2 == 0 else _C['row_odd'])
    c  = ws.cell(row=row, column=col, value=value)
    c.fill      = _fill(bg)
    c.font      = _font(bold=bold)
    c.alignment = _align(h=h, wrap=wrap)
    c.border    = _BORDER
    return c


# ─────────────────────────────  SHEET 1: Summary  ─────────────────────────────

def _build_summary_sheet(ws, snapshot: dict, plan_id: str):
    """
    Three sections:
      1. Plan-level inputs      (rows, cols, block_width, block_structure, broken_seats)
      2. Per-room configuration (from each room's own inputs block)
      3. Batch breakdown         (per room × per batch: counts, paper sets)
    """
    plan_inputs  = snapshot.get('inputs', {})
    rooms_data   = snapshot.get('rooms', {})
    room_names   = list(rooms_data.keys())

    r = 1   # Excel row pointer

    # ── 1. Plan-level inputs ──────────────────────────────────────────────────
    _section_title(ws, r, f'Plan Configuration  ·  {plan_id}', ncols=4)
    r += 1
    _col_header_row(ws, r, ['Parameter', 'Value'])
    r += 1

    plan_fields = [
        ('Rows per room',   plan_inputs.get('rows', '—')),
        ('Cols per room',   plan_inputs.get('cols', '—')),
        ('Block width',     plan_inputs.get('block_width', '—')),
        ('Block structure', ', '.join(str(x) for x in plan_inputs.get('block_structure', []))),
        ('Broken seats',    ', '.join(str(x) for x in plan_inputs.get('broken_seats', [])) or 'None'),
    ]
    for key, val in plan_fields:
        _data_cell(ws, r, 1, key, bold=True)
        _data_cell(ws, r, 2, str(val), h='center')
        r += 1

    r += 1  # blank gap

    # ── 2. Per-room configuration ─────────────────────────────────────────────
    _section_title(ws, r, 'Room Configurations', ncols=7)
    r += 1
    _col_header_row(ws, r, ['Room', 'Rows', 'Cols', 'Block Width',
                             'Block Structure', 'Broken Seats', 'Students Seated'])
    r += 1

    for room_name in room_names:
        room_info = rooms_data[room_name]
        ri        = room_info.get('inputs', {})
        bg        = _C['row_even'] if r % 2 == 0 else _C['row_odd']

        for ci, v in enumerate([
            room_name,
            ri.get('rows', '—'),
            ri.get('cols', '—'),
            ri.get('block_width', '—'),
            ', '.join(str(x) for x in ri.get('block_structure', [])),
            ', '.join(str(x) for x in ri.get('broken_seats', [])) or 'None',
            room_info.get('student_count', 0),
        ], 1):
            _data_cell(ws, r, ci, v, bg=bg,
                       h='center' if ci != 1 else 'left')
        r += 1

    r += 1  # blank gap

    # ── 3. Batch breakdown ────────────────────────────────────────────────────
    _section_title(ws, r, 'Batch Breakdown by Room', ncols=8)
    r += 1
    _col_header_row(ws, r, ['Room', 'Batch', 'Degree', 'Branch', 'Joining Year',
                             'Total Students', 'Paper Set A', 'Paper Set B'])
    r += 1

    grand_tot = grand_a = grand_b = 0
    for room_name in room_names:
        batches = rooms_data[room_name].get('batches', {})
        for batch_name, batch_data in batches.items():
            info     = batch_data.get('info', {})
            students = batch_data.get('students', [])
            cnt      = len(students)
            a_cnt    = sum(1 for s in students if s.get('paper_set') == 'A')
            b_cnt    = sum(1 for s in students if s.get('paper_set') == 'B')
            grand_tot += cnt; grand_a += a_cnt; grand_b += b_cnt
            bg = _C['row_even'] if r % 2 == 0 else _C['row_odd']
            for ci, v in enumerate([
                room_name, batch_name,
                info.get('degree', ''), info.get('branch', ''), info.get('joining_year', ''),
                cnt, a_cnt, b_cnt,
            ], 1):
                _data_cell(ws, r, ci, v, bg=bg,
                           h='center' if ci >= 6 else 'left')
            r += 1

    # Totals row
    for ci, v in enumerate(['TOTAL', '', '', '', '', grand_tot, grand_a, grand_b], 1):
        _data_cell(ws, r, ci, v, bg=_C['total_bg'], bold=True,
                   h='center' if ci >= 6 else 'left')

    ws.freeze_panes = 'A2'
    _auto_width(ws)


# ──────────────────────────────  SHEET PER ROOM  ──────────────────────────────

def _build_room_sheet(ws, room_name: str, room_info: dict):
    """
    Two sections in one sheet:
      ① Physical seating grid  — built from raw_matrix
         Rows = actual room rows (1-N),  Columns = seat-column letters (A, B, C…)
         Each cell: position / roll-number+set / student name  (3 lines, coloured)
      ② All-students detail table below the grid
    """
    raw_matrix    = room_info.get('raw_matrix', [])
    batches       = room_info.get('batches', {})
    room_inputs   = room_info.get('inputs', {})
    student_count = room_info.get('student_count', 0)

    n_rows = len(raw_matrix)
    n_cols = len(raw_matrix[0]) if raw_matrix else 0

    # ── Room header ───────────────────────────────────────────────────────────
    _section_title(ws, 1,
                   f'Room {room_name}  ·  {n_rows} rows × {n_cols} cols  '
                   f'·  {student_count} students seated',
                   ncols=max(n_cols + 1, 8))

    # Config info line
    config_text = (
        f"Block width: {room_inputs.get('block_width', '?')}  |  "
        f"Block structure: [{', '.join(str(x) for x in room_inputs.get('block_structure', []))}]  |  "
        f"Broken seats: {', '.join(str(x) for x in room_inputs.get('broken_seats', [])) or 'None'}"
    )
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=max(n_cols + 1, 8))
    cfg = ws.cell(row=2, column=1, value=config_text)
    cfg.font      = _font(size=9, colour='444444')
    cfg.alignment = _align(h='left')
    cfg.fill      = _fill(_C['section_bg'])

    # ── ① Seating Grid ────────────────────────────────────────────────────────
    GRID_HDR = 4   # Excel row of the column-letter header

    # Column-letter header
    hdr_lbl = ws.cell(row=GRID_HDR, column=1, value='↓ Row / Col →')
    hdr_lbl.font      = _font(bold=True, colour=_C['col_header_fg'], size=9)
    hdr_lbl.fill      = _fill(_C['col_header_bg'])
    hdr_lbl.alignment = _align(h='center')
    hdr_lbl.border    = _BORDER

    for ci in range(n_cols):
        c = ws.cell(row=GRID_HDR, column=ci + 2, value=_ALPHA[ci])
        c.font      = _font(bold=True, colour=_C['col_header_fg'])
        c.fill      = _fill(_C['col_header_bg'])
        c.alignment = _align(h='center')
        c.border    = _BORDER

    # Seat cells
    for ri, seat_row in enumerate(raw_matrix):
        xl_row = GRID_HDR + 1 + ri

        # Row-number label
        lbl = ws.cell(row=xl_row, column=1, value=f'Row {ri + 1}')
        lbl.font      = _font(bold=True, size=9, colour=_C['col_header_fg'])
        lbl.fill      = _fill(_C['col_header_bg'])
        lbl.alignment = _align(h='center')
        lbl.border    = _BORDER

        for ci, seat in enumerate(seat_row):
            xl_col = ci + 2
            is_broken  = bool(seat.get('is_broken', False))
            is_unalloc = bool(seat.get('is_unallocated', False))
            roll       = seat.get('roll_number') or ''
            name       = seat.get('student_name') or ''
            display    = seat.get('display', '') or ''
            pos        = seat.get('position', '')
            batch_hex  = (seat.get('color') or '#E9ECEF').lstrip('#')

            if is_broken:
                bg   = _C['broken']
                text = f'{pos}\nBROKEN'
                fg   = 'FFFFFF'
            elif is_unalloc or not roll:
                bg   = _C['unallocated']
                text = f'{pos}\n—'
                fg   = '888888'
            else:
                bg = batch_hex
                # pick contrasting text colour based on luminance
                try:
                    lum = (0.299 * int(bg[0:2], 16)
                           + 0.587 * int(bg[2:4], 16)
                           + 0.114 * int(bg[4:6], 16))
                    fg  = '111111' if lum > 150 else 'FFFFFF'
                except Exception:
                    fg = '111111'
                # 3-line display: position / roll+set / student name
                name_short = name[:18] if name else ''
                text = f'{pos}\n{display}\n{name_short}'

            cell = ws.cell(row=xl_row, column=xl_col, value=text)
            cell.fill      = _fill(bg)
            cell.font      = Font(color=fg, size=8)
            cell.alignment = _align(h='center', v='center', wrap=True)
            cell.border    = _BORDER

        # Taller rows for 3-line cell content
        ws.row_dimensions[xl_row].height = 46

    # Column widths for the grid
    ws.column_dimensions['A'].width = 12   # row-label column
    for ci in range(n_cols):
        ws.column_dimensions[get_column_letter(ci + 2)].width = 19

    grid_end_row = GRID_HDR + n_rows + 1

    # ── ② Full student detail table ───────────────────────────────────────────
    tbl_start = grid_end_row + 3

    _section_title(ws, tbl_start,
                   f'All Seated Students — Room {room_name}',
                   ncols=10)
    _col_header_row(ws, tbl_start + 1, [
        'Position', 'Roll Number', 'Student Name',
        'Batch', 'Degree', 'Branch', 'Joining Year',
        'Paper Set', 'Block', 'Status',
    ])

    tr = tbl_start + 2
    for batch_name, batch_data in batches.items():
        info     = batch_data.get('info', {})
        students = batch_data.get('students', [])
        for student in students:
            is_broken  = bool(student.get('is_broken', False))
            is_unalloc = bool(student.get('is_unallocated', False))
            status     = ('Broken'      if is_broken  else
                          'Unallocated' if is_unalloc else
                          'Seated')
            bg = (_C['broken']      if is_broken  else
                  _C['unallocated'] if is_unalloc else
                  (_C['row_even']   if tr % 2 == 0 else _C['row_odd']))

            for ci, v in enumerate([
                student.get('position', ''),
                student.get('roll_number', ''),
                student.get('student_name', ''),
                batch_name,
                info.get('degree', ''),
                info.get('branch', ''),
                info.get('joining_year', ''),
                student.get('paper_set', ''),
                student.get('block', ''),
                status,
            ], 1):
                _data_cell(ws, tr, ci, v, bg=bg,
                           h='center' if ci in (1, 8, 9) else 'left')
            tr += 1

    ws.freeze_panes = ws.cell(row=GRID_HDR + 1, column=2).coordinate


# ────────────────────────────  Workbook assembler  ────────────────────────────

def build_excel_workbook(snapshot: dict, plan_id: str) -> BytesIO:
    wb = Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    _build_summary_sheet(ws_summary, snapshot, plan_id)

    # One sheet per room
    rooms_data = snapshot.get('rooms', {})
    for room_name, room_info in rooms_data.items():
        sheet_name = f'Room_{room_name}'[:31]
        if sheet_name in wb.sheetnames:
            sheet_name = sheet_name[:28] + f'_{len(wb.sheetnames)}'
        ws_room = wb.create_sheet(title=sheet_name)
        _build_room_sheet(ws_room, room_name, room_info)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ──────────────────────────────────  API  ─────────────────────────────────────

def _verify_plan_ownership(plan_id: str, user_id: int) -> bool:
    try:
        from algo.database.db import get_db_connection
        conn = get_db_connection()
        cur  = conn.cursor()
        cur.execute("SELECT user_id FROM allocation_sessions WHERE plan_id = ?", (plan_id,))
        row  = cur.fetchone()
        conn.close()
        if not row:
            return True
        owner = row[0]
        return owner == user_id or owner is None
    except Exception as exc:
        logger.warning(f"Ownership check error: {exc}")
        return True


@excel_export_bp.route('/export-excel/<plan_id>', methods=['GET'])
@token_required
def export_excel(plan_id: str):
    """
    GET /api/export-excel/<plan_id>
    Returns a single .xlsx with:
      • Summary         — plan inputs, room configs, batch breakdown
      • Room_<X> sheets — physical seating grid + full student list per room
    """
    try:
        if not _verify_plan_ownership(plan_id, request.user_id):
            return jsonify({'error': 'Access denied — you do not own this plan'}), 403

        snapshot = _cache_mgr.load_snapshot(plan_id)
        if not snapshot:
            return jsonify({
                'error': 'Plan not found in cache',
                'hint':  'Re-generate the seating plan to refresh the cache.',
            }), 404

        excel_buf = build_excel_workbook(snapshot, plan_id)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename  = f'seating_{plan_id}_{timestamp}.xlsx'

        logger.info(f'✅ Excel export plan={plan_id} user={request.user_id}')

        return send_file(
            excel_buf,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    except Exception as exc:
        logger.error(f'Excel export error plan={plan_id}: {exc}', exc_info=True)
        return jsonify({'error': f'Export failed: {str(exc)}'}), 500

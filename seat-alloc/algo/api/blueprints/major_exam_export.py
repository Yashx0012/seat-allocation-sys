"""
Major Exam Export Blueprint
Handles Excel and other format exports
"""
from flask import Blueprint, request, jsonify, send_file
import io
import pandas as pd
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from algo.services.auth_service import token_required
from algo.core.cache.major_exam_cache import get_major_cache_manager

major_exam_export_bp = Blueprint('major_exam_export', __name__, url_prefix='/api/major-exam')

cache_manager = get_major_cache_manager()

_COLORS = {
    'sheet_header_bg': '1F4E79',
    'sheet_header_fg': 'FFFFFF',
    'col_header_bg': '2E75B6',
    'col_header_fg': 'FFFFFF',
    'row_even': 'F0F4FA',
    'row_odd': 'FFFFFF',
    'section_bg': 'DEEAF1',
    'total_bg': 'FFF3CD',
}

_THIN_SIDE = Side(style='thin', color='C7CBD1')
_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)


def _safe_int(value, default=0):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _clean_text(value):
    """Normalize values for clean export cells."""
    if value is None:
        return ''
    text = str(value).strip()
    if text.lower() in {'nan', 'none', 'null'}:
        return ''
    if text.endswith('.0'):
        number = text[:-2]
        if number.replace('-', '').isdigit():
            return number
    return text


def _append_students(rows, students, room_name, branch_name):
    """Append students with normalized room/branch values."""
    for student in students:
        row = dict(student)
        row['room_name'] = room_name
        row['branch'] = row.get('branch') or branch_name
        rows.append(row)


def _fill(hex_color: str):
    color = hex_color.lstrip('#')
    return PatternFill(start_color=color, end_color=color, fill_type='solid')


def _font(bold=False, color='000000', size=10):
    return Font(bold=bold, color=color.lstrip('#'), size=size)


def _align(horizontal='left', vertical='center', wrap=False):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


def _auto_width(ws, padding=2, max_width=52):
    from openpyxl.cell.cell import MergedCell

    for col in ws.columns:
        best = 0
        col_letter = None
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            value = '' if cell.value is None else str(cell.value)
            if len(value) > best:
                best = len(value)
        if col_letter:
            ws.column_dimensions[col_letter].width = min(best + padding, max_width)


def _style_table_sheet(ws):
    if ws.max_row < 1:
        return

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _fill(_COLORS['col_header_bg'])
        cell.font = _font(bold=True, color=_COLORS['col_header_fg'])
        cell.alignment = _align(horizontal='center')
        cell.border = _BORDER

    for row in range(2, ws.max_row + 1):
        bg = _COLORS['row_even'] if row % 2 == 0 else _COLORS['row_odd']
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = _fill(bg)
            cell.border = _BORDER
            if isinstance(cell.value, (int, float)):
                cell.alignment = _align(horizontal='center')
            else:
                cell.alignment = _align(horizontal='left')

    ws.freeze_panes = 'A2'
    _auto_width(ws)


def _section_title(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = _fill(_COLORS['sheet_header_bg'])
    cell.font = _font(bold=True, color=_COLORS['sheet_header_fg'], size=11)
    cell.alignment = _align(horizontal='left')
    cell.border = _BORDER


def _header_row(ws, row, labels):
    for col_idx, label in enumerate(labels, 1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.fill = _fill(_COLORS['col_header_bg'])
        cell.font = _font(bold=True, color=_COLORS['col_header_fg'])
        cell.alignment = _align(horizontal='center')
        cell.border = _BORDER


def _data_row(ws, row, values, alignments=None, bold=False, bg=None):
    fill_color = bg or (_COLORS['row_even'] if row % 2 == 0 else _COLORS['row_odd'])
    for col_idx, value in enumerate(values, 1):
        align_h = 'left'
        if alignments and col_idx - 1 < len(alignments):
            align_h = alignments[col_idx - 1]
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.fill = _fill(fill_color)
        cell.font = _font(bold=bold)
        cell.alignment = _align(horizontal=align_h)
        cell.border = _BORDER


def _room_rows(plan_data: dict):
    rows = []
    rooms = plan_data.get('rooms', [])

    def push_room(room_name, room_data):
        if not isinstance(room_data, dict):
            return
        branch_allocs = room_data.get('branch_allocations', {})
        batches = room_data.get('batches', {})

        if isinstance(branch_allocs, dict) and branch_allocs:
            branch_count = len(branch_allocs)
            total_students = _safe_int(
                room_data.get('total_students'),
                sum(_safe_int(a.get('count'), len(a.get('students', []))) for a in branch_allocs.values() if isinstance(a, dict))
            )
        elif isinstance(batches, dict) and batches:
            branch_count = len(batches)
            total_students = _safe_int(
                room_data.get('total_students') or room_data.get('allocated_count'),
                sum(len((b.get('students') or [])) for b in batches.values() if isinstance(b, dict))
            )
        else:
            branch_count = 0
            total_students = _safe_int(room_data.get('total_students') or room_data.get('allocated_count'), 0)

        rows.append({
            'room_name': _clean_text(room_name or room_data.get('room_name') or room_data.get('name')),
            'capacity': _safe_int(room_data.get('capacity'), 0),
            'total_students': total_students,
            'branch_count': branch_count,
        })

    if isinstance(rooms, list):
        for room in rooms:
            if isinstance(room, dict):
                push_room(room.get('room_name') or room.get('name'), room)
    elif isinstance(rooms, dict):
        for room_name, room_data in rooms.items():
            push_room(room_name, room_data)

    return sorted(rows, key=lambda r: r['room_name'])


def _branch_rows(students, metadata):
    branch_map = {}

    for student in students:
        branch = _clean_text(student.get('branch')) or 'UNSPECIFIED'
        room = _clean_text(student.get('room_name')) or '-'
        if branch not in branch_map:
            branch_map[branch] = {'count': 0, 'rooms': set()}
        branch_map[branch]['count'] += 1
        if room and room != '-':
            branch_map[branch]['rooms'].add(room)

    if not branch_map:
        counts = metadata.get('branch_counts', {}) if isinstance(metadata.get('branch_counts', {}), dict) else {}
        for branch, count in counts.items():
            branch_map[_clean_text(branch) or 'UNSPECIFIED'] = {'count': _safe_int(count, 0), 'rooms': set()}

    rows = []
    for branch, data in branch_map.items():
        rows.append({
            'branch': branch,
            'count': data['count'],
            'rooms': len(data['rooms']),
        })

    rows.sort(key=lambda r: r['branch'])
    return rows


def _build_summary_sheet(ws, plan_data: dict, students):
    metadata = plan_data.get('metadata', {}) if isinstance(plan_data.get('metadata', {}), dict) else {}
    rooms = _room_rows(plan_data)
    branches = _branch_rows(students, metadata)

    total_students = metadata.get('total_students', len(students))
    allocated_count = metadata.get('allocated_count', len(students))
    room_count = metadata.get('room_count', len(rooms))
    exported_at = datetime.utcnow().isoformat() + 'Z'

    row = 1

    _section_title(ws, row, f"Major Exam Plan Summary  ·  {plan_data.get('plan_id', '-')}", 4)
    row += 1
    _header_row(ws, row, ['Parameter', 'Value'])
    row += 1

    overview_rows = [
        ('Plan ID', plan_data.get('plan_id', '-')),
        ('Status', metadata.get('status', 'pending')),
        ('Total Students', _safe_int(total_students, len(students))),
        ('Allocated Students', _safe_int(allocated_count, len(students))),
        ('Room Count', _safe_int(room_count, len(rooms))),
        ('Created At', plan_data.get('created_at', '-')),
        ('Finalized At', metadata.get('finalized_at', '-')),
        ('Exported At (UTC)', exported_at),
    ]

    for key, value in overview_rows:
        is_numeric = isinstance(value, int)
        _data_row(ws, row, [key, value], alignments=['left', 'center' if is_numeric else 'left'])
        row += 1

    row += 1
    _section_title(ws, row, 'Room Configurations', 4)
    row += 1
    _header_row(ws, row, ['Room', 'Capacity', 'Allocated Students', 'Branch Count'])
    row += 1

    if rooms:
        for room_info in rooms:
            _data_row(
                ws,
                row,
                [room_info['room_name'], room_info['capacity'], room_info['total_students'], room_info['branch_count']],
                alignments=['left', 'center', 'center', 'center']
            )
            row += 1
    else:
        _data_row(ws, row, ['No room allocation data available', '', '', ''], alignments=['left', 'left', 'left', 'left'])
        row += 1

    row += 1
    _section_title(ws, row, 'Branch Breakdown', 4)
    row += 1
    _header_row(ws, row, ['Branch', 'Students', 'Assigned Rooms', 'Share (%)'])
    row += 1

    total_branch_students = sum(r['count'] for r in branches) if branches else 0
    if branches:
        for branch_info in branches:
            share = round((branch_info['count'] / total_branch_students) * 100, 2) if total_branch_students else 0
            _data_row(
                ws,
                row,
                [branch_info['branch'], branch_info['count'], branch_info['rooms'], share],
                alignments=['left', 'center', 'center', 'center']
            )
            row += 1

        _data_row(
            ws,
            row,
            ['TOTAL', total_branch_students, sum(r['rooms'] for r in branches), 100 if total_branch_students else 0],
            alignments=['left', 'center', 'center', 'center'],
            bold=True,
            bg=_COLORS['total_bg']
        )
    else:
        _data_row(ws, row, ['No branch data available', '', '', ''], alignments=['left', 'left', 'left', 'left'])

    ws.freeze_panes = 'A2'
    _auto_width(ws)


def _flatten_students(plan_data: dict):
    """Flatten students from new + legacy room/branch/flat shapes."""
    rows = []

    rooms = plan_data.get('rooms', [])
    if isinstance(rooms, list) and rooms:
        for room in rooms:
            if not isinstance(room, dict):
                continue

            room_name = room.get('room_name') or room.get('name') or ''
            branch_allocs = room.get('branch_allocations', {})
            if isinstance(branch_allocs, dict) and branch_allocs:
                for branch_name, alloc in branch_allocs.items():
                    if isinstance(alloc, dict):
                        _append_students(rows, alloc.get('students', []), room_name, branch_name)
                continue

            # Legacy list room schema with batches
            batches = room.get('batches', {})
            if isinstance(batches, dict):
                for batch_name, batch in batches.items():
                    if not isinstance(batch, dict):
                        continue
                    info = batch.get('info', {}) if isinstance(batch.get('info', {}), dict) else {}
                    branch_name = info.get('branch') or batch_name
                    _append_students(rows, batch.get('students', []), room_name, branch_name)
        return rows

    # Legacy dict room schema keyed by room name
    if isinstance(rooms, dict) and rooms:
        for room_name, room_data in rooms.items():
            if not isinstance(room_data, dict):
                continue

            branch_allocs = room_data.get('branch_allocations', {})
            if isinstance(branch_allocs, dict) and branch_allocs:
                for branch_name, alloc in branch_allocs.items():
                    if isinstance(alloc, dict):
                        _append_students(rows, alloc.get('students', []), room_name, branch_name)
                continue

            batches = room_data.get('batches', {})
            if isinstance(batches, dict):
                for batch_name, batch in batches.items():
                    if not isinstance(batch, dict):
                        continue
                    info = batch.get('info', {}) if isinstance(batch.get('info', {}), dict) else {}
                    branch_name = info.get('branch') or batch_name
                    _append_students(rows, batch.get('students', []), room_name, branch_name)
        if rows:
            return rows

    branches = plan_data.get('branches', {})
    if isinstance(branches, dict) and branches:
        for branch_name, branch_students in branches.items():
            for student in branch_students:
                row = dict(student)
                row['branch'] = row.get('branch') or branch_name
                rows.append(row)
        return rows

    legacy = plan_data.get('students', [])
    if isinstance(legacy, list):
        return [dict(s) for s in legacy]

    return []


def generate_excel_export(plan_data: dict) -> bytes:
    """Generate a clean Excel file from processed plan data."""
    try:
        output = io.BytesIO()

        students = _flatten_students(plan_data)
        students = sorted(
            students,
            key=lambda s: (
                _clean_text(s.get('room_name', '')),
                _clean_text(s.get('branch', '')),
                _clean_text(s.get('enrollment', '')),
            )
        )

        processed_rows = []
        for idx, student in enumerate(students, start=1):
            processed_rows.append({
                'S.No': idx,
                'Enrollment No': _clean_text(student.get('enrollment')),
                'Student Name': _clean_text(student.get('name')),
                'Branch': _clean_text(student.get('branch')),
                'Room': _clean_text(student.get('room_name')),
                'Code': _clean_text(student.get('code')),
                'Password': _clean_text(student.get('password')),
            })

        df = pd.DataFrame(
            processed_rows,
            columns=['S.No', 'Enrollment No', 'Student Name', 'Branch', 'Room', 'Code', 'Password']
        )
        if df.empty:
            df = pd.DataFrame(columns=['S.No', 'Enrollment No', 'Student Name', 'Branch', 'Room', 'Code', 'Password'])

        if df.empty:
            room_summary_df = pd.DataFrame(columns=['Room', 'Branch', 'Count'])
        else:
            room_summary_df = (
                df.groupby(['Room', 'Branch'], dropna=False)
                .size()
                .reset_index(name='Count')
                .sort_values(by=['Room', 'Branch'])
            )
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Students', index=False)
            room_summary_df.to_excel(writer, sheet_name='Room Summary', index=False)

            workbook = writer.book
            summary_ws = workbook.create_sheet('Summary', 0)
            _build_summary_sheet(summary_ws, plan_data, students)

            students_ws = writer.sheets.get('Students')
            if students_ws:
                _style_table_sheet(students_ws)

            room_summary_ws = writer.sheets.get('Room Summary')
            if room_summary_ws:
                _style_table_sheet(room_summary_ws)
        
        output.seek(0)
        return output.getvalue()
    
    except Exception as e:
        print(f"❌ Error generating Excel: {e}")
        return None


@major_exam_export_bp.route('/download/excel/<plan_id>', methods=['GET'])
@token_required
def download_excel(plan_id):
    """Download student data as Excel"""
    try:
        user_id = request.user_id
        
        plan = cache_manager.retrieve_plan(user_id, plan_id)
        if not plan:
            return jsonify({'error': 'Plan not found'}), 404
        
        excel_data = generate_excel_export(plan)
        if not excel_data:
            return jsonify({'error': 'Failed to generate Excel file'}), 500
        
        return send_file(
            io.BytesIO(excel_data),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'major_exam_{plan_id}_students.xlsx'
        )
    
    except Exception as e:
        print(f"❌ Error downloading Excel: {e}")
        return jsonify({'error': f'Server error: {str(e)}'}), 500


@major_exam_export_bp.route('/download/json/<plan_id>', methods=['GET'])
@token_required
def download_json(plan_id):
    """Download plan as JSON"""
    try:
        user_id = request.user_id
        
        plan = cache_manager.retrieve_plan(user_id, plan_id)
        if not plan:
            return jsonify({'error': 'Plan not found'}), 404
        
        # Prepare response
        response_data = io.BytesIO()
        import json
        json_str = json.dumps(plan, indent=2, default=str)
        response_data.write(json_str.encode())
        response_data.seek(0)
        
        return send_file(
            response_data,
            mimetype='application/json',
            as_attachment=True,
            download_name=f'major_exam_{plan_id}.json'
        )
    
    except Exception as e:
        print(f"❌ Error downloading JSON: {e}")
        return jsonify({'error': f'Server error: {str(e)}'}), 500
